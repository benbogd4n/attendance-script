# install pyinstaller to be able to convert it into a .exe
# pip install pyinstaller

# Install pywhatkit using pip to be able to run script:
# pip install pywhatkit

from openpyxl import load_workbook
import datetime
import pywhatkit

# don't change variables
now = datetime.datetime.now()
str1 = "Check out at "
myName = input("What's your name? ").strip().capitalize()
formatted_minute = str(now.minute).zfill(2)

# var for whatsapp group id
group_id = "KgjLth6nChq0ci1aj3WgGR"


# test message to be sent
print(f"{str1}{now.hour}:{formatted_minute}")

# defines group id, message to send
# script will require to be logged in on web whatsapp and will take a few seconds to run
# opens a new browser tab and types the msg in the checkin chat

pywhatkit.sendwhatmsg_to_group_instantly(group_id, f"{str1}{now.hour}:{formatted_minute}")

# Define the checkInInfo function here
def checkInInfo():
    x = {
        "name": myName,
        "checkOut": f"{now.hour}:{formatted_minute}"
    }
    return x

# Don't change variables
checkInOutput = checkInInfo()

# Load existing workbook
wb = load_workbook('CheckInList.xlsx')
ws = wb.active
empty_row = ws.max_row

print(checkInOutput)

# Search for the name in column B
name_column = ws['B']
target_row = None

for cell in name_column:
    if cell.value == checkInOutput["name"]:
        target_row = cell.row
        break

# If the name was found, find the first cell with "-" in the same row
if target_row:
    empty_column = 3

    # Find the first cell with "." in the same row as the target name
    for col in range(3, ws.max_column + 1):
        if ws.cell(row=target_row, column=col).value == "-":
            empty_column = col
            break

    # Overwrite the cell with the new checkOut value
    ws.cell(row=target_row, column=empty_column, value=checkInOutput["checkOut"])

wb.save('CheckInList.xlsx')
