# install pywhatkit using pip to be able to run script:
# pip install pywhatkit

# install pyinstaller to be able to convert it into a .exe
# pip install pyinstaller

import pywhatkit
from openpyxl import Workbook, load_workbook
import datetime

# don't change variables
now = datetime.datetime.now()
str1 = "Check in at "
myName = input("What's your name? ").strip().capitalize()
myStatus = input("What's your current work status? ").strip().capitalize()
formatted_minute = str(now.minute).zfill(2)

# var for whatsapp group id
group_id = "KgjLth6nChq0ci1aj3WgGR"

# test message to be sent
print(f"{str1}{now.hour}:{formatted_minute} - {myStatus}")

# defines group id, message to send
# script will require to be logged in on web whatsapp and will take a few seconds to run
# opens a new browser tab and types the msg in the checkin chat
pywhatkit.sendwhatmsg_to_group_instantly(group_id, f"{str1}{now.hour}:{formatted_minute} - {myStatus}")

def checkInInfo():
    now = datetime.datetime.now()

    x = {
        "date": f"{now.day}-{now.month}-{now.year}",
        "name": myName,
        "status": myStatus,
        "checkIn": f"{now.hour}:{formatted_minute}"
    }

    return [x]

checkInOutput = checkInInfo()

# load existing workbook
wb = load_workbook('CheckInList.xlsx')
ws = wb.active
empty_row = ws.max_row

print(checkInOutput)

# Search for the name in column B
name_column = ws['B']
target_row = None

for cell in name_column:
    if cell.value == myName:
        target_row = cell.row
        break

# If the name was found, find the last filled cell in the same row
if target_row:
    empty_column = 3

    # Find the last filled cell in the same row as the target name
    last_filled_column = None
    for col in range(3, ws.max_column + 1):
        if ws.cell(row=target_row, column=col).value:
            last_filled_column = col

    # If the last filled cell was found, insert the new entry after it
    if last_filled_column:
        empty_column = last_filled_column + 1

    for entry in checkInOutput:
        ws.cell(row=target_row, column=empty_column, value=entry["date"])
        ws.cell(row=target_row, column=empty_column + 1, value=entry["status"])
        ws.cell(row=target_row, column=empty_column + 2, value=entry["checkIn"])
        ws.cell(row=target_row, column=empty_column + 3, value="-")

wb.save('CheckInList.xlsx')